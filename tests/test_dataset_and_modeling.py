from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import numpy as np
import pandas as pd

from pipeline.dataset import FactorDatasetBuilder, PreparedData
from pipeline.modeling import train_dual_regime_models


class DatasetAndModelingTest(unittest.TestCase):
    def test_read_raw_data_retries_after_timeout(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            raw_path = root / "RBZL.SHF.csv"
            merged_cache = root / "cache.parquet"
            cache_meta = root / "cache_meta.json"
            regime_plot = root / "regime_plot.png"

            builder = FactorDatasetBuilder(
                config_override={
                    "paths": {
                        "raw_data": str(raw_path),
                        "merged_cache": str(merged_cache),
                        "cache_meta": str(cache_meta),
                        "regime_plot": str(regime_plot),
                    },
                }
            )

            sample_df = pd.DataFrame(
                {
                    "TDATE": ["2024-01-01 09:00:00"],
                    "OPEN": [1.0],
                    "HIGH": [1.2],
                    "LOW": [0.9],
                    "CLOSE": [1.1],
                }
            )

            with patch("pipeline.dataset.pd.read_csv", side_effect=[TimeoutError(60, "Operation timed out"), sample_df]) as mocked_read_csv:
                raw = builder._read_raw_data()

            self.assertEqual(mocked_read_csv.call_count, 2)
            self.assertEqual(len(raw), 1)
            self.assertEqual(raw["CLOSE"].iloc[0], 1.1)

    def test_mid_weekly_merge_uses_asof_forward_fill(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            raw_path = root / "RBZL.SHF.csv"
            mid_dir = root / "mid_weekly"
            mid_dir.mkdir()
            merged_cache = root / "cache.parquet"
            cache_meta = root / "cache_meta.json"
            regime_plot = root / "regime_plot.png"

            raw_df = pd.DataFrame(
                {
                    "TDATE": pd.to_datetime(
                        [
                            "2024-01-01 09:00:00",
                            "2024-01-02 09:00:00",
                            "2024-01-03 09:00:00",
                        ]
                    ),
                    "OPEN": [1, 2, 3],
                    "HIGH": [2, 3, 4],
                    "LOW": [1, 2, 3],
                    "CLOSE": [1.5, 2.5, 3.5],
                    "VOLUME": [10, 20, 30],
                    "AMOUNT": [15, 50, 105],
                }
            )
            raw_df.to_csv(raw_path, index=False)

            pd.DataFrame(
                {
                    "date": ["2024-01-01", "2024-01-03"],
                    "inventory": [100, 200],
                }
            ).to_csv(mid_dir / "inventory.csv", index=False)

            builder = FactorDatasetBuilder(
                config_override={
                    "paths": {
                        "raw_data": str(raw_path),
                        "merged_cache": str(merged_cache),
                        "cache_meta": str(cache_meta),
                        "regime_plot": str(regime_plot),
                        "mid_weekly_dir": str(mid_dir),
                    },
                    "data": {
                        "use_mid_weekly": True,
                    },
                    "product": {
                        "product_id": "RB",
                        "mid_weekly_files": ["inventory.csv"],
                    },
                }
            )
            raw = builder._read_raw_data()
            merged, mid_cols = builder._merge_mid_weekly_features(raw)

            # level col + AVAILABLE dummy in that order (derived disabled via < 4 obs).
            self.assertEqual(mid_cols, ["MID_inventory", "MID_inventory_AVAILABLE"])
            self.assertEqual(merged["MID_inventory"].tolist(), [100.0, 100.0, 200.0])
            self.assertEqual(merged["MID_inventory_AVAILABLE"].tolist(), [1, 1, 1])

    def test_mid_weekly_xlsx_wide_format(self) -> None:
        """T3.1: xlsx wide-format reader. 4-row header + multi-indicator + NaN tail."""
        import openpyxl

        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            raw_path = root / "RBZL.SHF.csv"
            mid_dir = root / "mid_weekly"
            mid_dir.mkdir()
            merged_cache = root / "cache.parquet"
            cache_meta = root / "cache_meta.json"
            regime_plot = root / "regime_plot.png"

            # 5-min raw bars spanning 3 calendar weeks (2024-01-01 Mon .. 2024-01-21 Sun).
            raw_index = pd.date_range("2024-01-01 09:00:00", "2024-01-19 14:55:00", freq="5min")
            raw_index = raw_index[(raw_index.hour >= 9) & (raw_index.hour < 15) & (raw_index.weekday < 5)]
            raw_df = pd.DataFrame(
                {
                    "TDATE": raw_index,
                    "OPEN": np.arange(len(raw_index), dtype=float) + 1.0,
                    "HIGH": np.arange(len(raw_index), dtype=float) + 2.0,
                    "LOW": np.arange(len(raw_index), dtype=float),
                    "CLOSE": np.arange(len(raw_index), dtype=float) + 1.5,
                    "VOLUME": 10,
                    "AMOUNT": 100,
                }
            )
            raw_df.to_csv(raw_path, index=False)

            # Synthesize a 2-indicator xlsx with the strict 4-row header convention.
            # indA: has weekly values on all 3 Mondays.
            # indB: valid only on 2024-01-01 Mon and 2024-01-08 Mon; NaN on 2024-01-15.
            xlsx_path = mid_dir / "RB.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="单位")
            ws.cell(row=1, column=2, value="万吨")
            ws.cell(row=1, column=3, value="%")
            ws.cell(row=2, column=1, value="指标名称")
            ws.cell(row=2, column=2, value="螺纹钢社会库存")
            ws.cell(row=2, column=3, value="高炉开工率")
            ws.cell(row=3, column=1, value="频率")
            ws.cell(row=3, column=2, value="周")
            ws.cell(row=3, column=3, value="周")
            ws.cell(row=4, column=1, value="指标ID")
            ws.cell(row=4, column=2, value="S123456")
            ws.cell(row=4, column=3, value="S789012")

            ws.cell(row=5, column=1, value=pd.Timestamp("2024-01-15"))
            ws.cell(row=5, column=2, value=300.0)
            # col 3 (indB) intentionally blank on 2024-01-15
            ws.cell(row=6, column=1, value=pd.Timestamp("2024-01-08"))
            ws.cell(row=6, column=2, value=280.0)
            ws.cell(row=6, column=3, value=82.0)
            ws.cell(row=7, column=1, value=pd.Timestamp("2024-01-01"))
            ws.cell(row=7, column=2, value=260.0)
            ws.cell(row=7, column=3, value=80.0)
            wb.save(xlsx_path)

            builder = FactorDatasetBuilder(
                config_override={
                    "paths": {
                        "raw_data": str(raw_path),
                        "merged_cache": str(merged_cache),
                        "cache_meta": str(cache_meta),
                        "regime_plot": str(regime_plot),
                        "mid_weekly_dir": str(mid_dir),
                    },
                    "data": {"use_mid_weekly": True},
                    "product": {"product_id": "RB", "mid_weekly_files": ["RB.xlsx"]},
                }
            )
            raw = builder._read_raw_data()
            merged, mid_cols = builder._merge_mid_weekly_features(raw)

            # Two level MID_* columns exist and carry the product prefix.
            level_cols = [c for c in mid_cols if not c.endswith("_AVAILABLE") and "_RET_" not in c and "_ZSCORE_" not in c and "_PCT_RANK_" not in c]
            self.assertEqual(len(level_cols), 2)
            for col in level_cols:
                self.assertTrue(col.startswith("MID_RB_"), f"column {col} missing MID_RB_ prefix")
            # Column-level metadata persisted with frequency / indicator_id.
            self.assertEqual(len(builder._mid_weekly_metadata), 2)
            for meta in builder._mid_weekly_metadata.values():
                self.assertEqual(meta["frequency"], "周")
                self.assertIn(meta["indicator_id"], {"S123456", "S789012"})

            # Timestamp alignment: value at the Monday bar == that Monday's xlsx value
            # (backward merge_asof picks the same-day record).
            ind_a = next(c for c in level_cols if builder._mid_weekly_metadata[c]["indicator_id"] == "S123456")
            ind_b = next(c for c in level_cols if builder._mid_weekly_metadata[c]["indicator_id"] == "S789012")

            def at(ts: str) -> pd.Series:
                row = merged.loc[merged["TDATE"] == pd.Timestamp(ts)]
                self.assertEqual(len(row), 1, f"no row at {ts}")
                return row.iloc[0]

            self.assertAlmostEqual(float(at("2024-01-01 09:00:00")[ind_a]), 260.0, places=3)
            self.assertAlmostEqual(float(at("2024-01-08 09:00:00")[ind_a]), 280.0, places=3)
            self.assertAlmostEqual(float(at("2024-01-15 09:00:00")[ind_a]), 300.0, places=3)

            # Within-week forward fill: Friday of the first week should still carry
            # Monday's value (no bfill → no future leakage).
            self.assertAlmostEqual(float(at("2024-01-05 14:55:00")[ind_a]), 260.0, places=3)

            # No future leakage: the Tuesday of week 1 (2024-01-02) predates indB's
            # first valid date (2024-01-01 09:00 Mon) — wait, 01-01 is Mon so
            # Tue 01-02 is after. Pick an earlier bar: we chose raw start 01-01 09:00,
            # so the very first raw bar IS the xlsx timestamp. Confirm ffill for
            # Tue carries Mon's indB value:
            self.assertAlmostEqual(float(at("2024-01-02 10:00:00")[ind_b]), 80.0, places=3)
            # After indB goes NaN (Mon 01-15), merge_asof + ffill should carry the
            # last known value (82.0 from 01-08), not leak the future. The xlsx
            # row for 01-15 has indB blank, so the aligned value stays 82.0.
            self.assertAlmostEqual(float(at("2024-01-15 09:00:00")[ind_b]), 82.0, places=3)
            self.assertAlmostEqual(float(at("2024-01-19 14:55:00")[ind_b]), 82.0, places=3)

    def test_mid_weekly_strategy_b_available_clamp_and_derived(self) -> None:
        """T3.2: MID_*_AVAILABLE dummy, ffill_max_bars clamp, derived RET/ZSCORE/PCT_RANK."""
        import openpyxl

        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            raw_path = root / "RBZL.SHF.csv"
            mid_dir = root / "mid_weekly"
            mid_dir.mkdir()
            merged_cache = root / "cache.parquet"
            cache_meta = root / "cache_meta.json"
            regime_plot = root / "regime_plot.png"

            # 6 weekly Mondays' worth of 5-min bars (so derived window=4 can emit).
            raw_index = pd.date_range("2024-01-01 09:00:00", "2024-02-09 14:55:00", freq="5min")
            raw_index = raw_index[(raw_index.hour >= 9) & (raw_index.hour < 15) & (raw_index.weekday < 5)]
            raw_df = pd.DataFrame({
                "TDATE": raw_index,
                "OPEN": np.arange(len(raw_index), dtype=float) + 1.0,
                "HIGH": np.arange(len(raw_index), dtype=float) + 2.0,
                "LOW": np.arange(len(raw_index), dtype=float),
                "CLOSE": np.arange(len(raw_index), dtype=float) + 1.5,
                "VOLUME": 10,
                "AMOUNT": 100,
            })
            raw_df.to_csv(raw_path, index=False)

            # xlsx with 1 indicator, 6 observations, all present.
            xlsx_path = mid_dir / "RB.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="单位"); ws.cell(row=1, column=2, value="万吨")
            ws.cell(row=2, column=1, value="指标名称"); ws.cell(row=2, column=2, value="螺纹钢社会库存")
            ws.cell(row=3, column=1, value="频率"); ws.cell(row=3, column=2, value="周")
            ws.cell(row=4, column=1, value="指标ID"); ws.cell(row=4, column=2, value="S999001")
            # Descending dates matching real xlsx layout.
            for row_idx, (date_str, value) in enumerate([
                ("2024-02-05", 320.0),  # Mon week 6
                ("2024-01-29", 315.0),
                ("2024-01-22", 310.0),
                ("2024-01-15", 305.0),
                ("2024-01-08", 300.0),
                ("2024-01-01", 290.0),
            ], start=5):
                ws.cell(row=row_idx, column=1, value=pd.Timestamp(date_str))
                ws.cell(row=row_idx, column=2, value=value)
            wb.save(xlsx_path)

            # Tight clamp: 20 5-min bars of staleness max — forces NaN shortly after each Monday.
            builder = FactorDatasetBuilder(
                config_override={
                    "paths": {
                        "raw_data": str(raw_path),
                        "merged_cache": str(merged_cache),
                        "cache_meta": str(cache_meta),
                        "regime_plot": str(regime_plot),
                        "mid_weekly_dir": str(mid_dir),
                    },
                    "data": {"use_mid_weekly": True},
                    "product": {"product_id": "RB", "mid_weekly_files": ["RB.xlsx"]},
                    "mid_weekly": {
                        "available_dummy": True,
                        "ffill_max_bars": 20,
                        "level_keep": True,
                        "derived": {
                            "enabled": True,
                            "rolling_windows": [4],
                            "transforms": ["ret", "zscore", "pct_rank"],
                        },
                    },
                }
            )
            raw = builder._read_raw_data()
            merged, mid_cols = builder._merge_mid_weekly_features(raw)

            # Bucket the columns.
            level = [c for c in mid_cols if not c.endswith("_AVAILABLE") and all(s not in c for s in ("_RET_", "_ZSCORE_", "_PCT_RANK_"))]
            avail = [c for c in mid_cols if c.endswith("_AVAILABLE")]
            derived = [c for c in mid_cols if any(s in c for s in ("_RET_", "_ZSCORE_", "_PCT_RANK_"))]

            self.assertEqual(len(level), 1, f"expected 1 level col, got {level}")
            self.assertEqual(len(avail), 1, f"expected 1 AVAILABLE col, got {avail}")
            # window=4 × 3 transforms = 3 derived cols for the single indicator.
            self.assertEqual(len(derived), 3, f"expected 3 derived cols, got {derived}")

            level_col = level[0]
            avail_col = avail[0]

            # Clamp correctness: Monday 01-08 09:00 — brand-new arrival → level not NaN,
            # AVAILABLE=1. Bar roughly 21 5-min steps later on the same day must be
            # NaN'd out (staleness > ffill_max_bars=20) and AVAILABLE must flip to 0.
            mon_idx = merged.index[merged["TDATE"] == pd.Timestamp("2024-01-08 09:00:00")][0]
            self.assertFalse(pd.isna(merged.loc[mon_idx, level_col]))
            self.assertEqual(int(merged.loc[mon_idx, avail_col]), 1)
            stale_idx = merged.index[merged["TDATE"] == pd.Timestamp("2024-01-08 11:00:00")][0]
            # 21 bars after 09:00 on a trading day = 11:05 given 5-min bars but since we
            # built raw_df with hour-range filter, the actual bar indexing starts at 09:00.
            # Pick a later bar to be safe: after 25 bars we're definitely stale.
            later_idx = mon_idx + 25
            self.assertTrue(pd.isna(merged.loc[later_idx, level_col]))
            self.assertEqual(int(merged.loc[later_idx, avail_col]), 0)

            # Derived factor values after window=4 observations exist. The 5th Monday
            # (2024-01-29) gives pct_change(4) = 315/290 - 1 ≈ 0.0862.
            ret_col = next(c for c in derived if c.endswith("_RET_4"))
            mon5_idx = merged.index[merged["TDATE"] == pd.Timestamp("2024-01-29 09:00:00")][0]
            self.assertAlmostEqual(float(merged.loc[mon5_idx, ret_col]), 315.0 / 290.0 - 1.0, places=4)

    def test_train_dual_regime_models_skips_model_persistence(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            model_dir = root / "models"
            summary_path = root / "training_summary.json"
            training_plot = root / "training_plot.png"
            comparison_plot = root / "comparison_plot.png"

            def _make_split(start: str, regime_label: int, n: int) -> pd.DataFrame:
                idx = pd.date_range(start, periods=n, freq="min")
                base = np.linspace(-1.0, 1.0, n) + regime_label * 0.2
                future_return = base * 0.01
                return pd.DataFrame(
                    {
                        "TDATE": idx,
                        "TRADE_DATE": idx.normalize(),
                        "REGIME_LABEL": regime_label,
                        "future_return": future_return,
                        "target_vol_scale": np.ones(n),
                        "target_vol_norm": future_return,
                        "f1": base,
                    }
                )

            train_df = pd.concat([_make_split("2024-01-01 09:00:00", -1, 24), _make_split("2024-01-02 09:00:00", 1, 24)], ignore_index=True)
            val_df = pd.concat([_make_split("2024-01-03 09:00:00", -1, 12), _make_split("2024-01-04 09:00:00", 1, 12)], ignore_index=True)
            test_df = pd.concat([_make_split("2024-01-05 09:00:00", -1, 12), _make_split("2024-01-06 09:00:00", 1, 12)], ignore_index=True)
            full_df = pd.concat([train_df, val_df, test_df], ignore_index=True)

            prepared = PreparedData(
                full_data=full_df,
                train_data=train_df,
                val_data=val_df,
                test_data=test_df,
                feature_cols=["f1"],
                factor_cols=["f1"],
                engineered_cols=[],
                runtime_factor_cols=["f1"],
                mid_weekly_cols=[],
                target_col="target_vol_norm",
                metadata={"feature_count": 1},
                feature_manifest={"all_feature_cols": ["f1"]},
            )

            _, summary, _ = train_dual_regime_models(
                prepared=prepared,
                config_override={
                    "paths": {
                        "model_dir": str(model_dir),
                        "training_summary": str(summary_path),
                        "training_plot": str(training_plot),
                        "training_comparison_plot": str(comparison_plot),
                    },
                    "model": {
                        "persist_models": False,
                        "num_boost_round": 10,
                        "early_stopping_rounds": 0,
                        "feature_importance_top_n": 5,
                        "common_params": {
                            "objective": "regression",
                            "metric": "l2",
                            "boosting_type": "gbdt",
                            "learning_rate": 0.1,
                            "num_leaves": 7,
                            "max_depth": 3,
                            "min_child_samples": 2,
                            "verbosity": -1,
                            "seed": 42,
                        },
                        "low_vol_overrides": {},
                        "high_vol_overrides": {},
                    },
                },
            )

            self.assertFalse(model_dir.exists())
            self.assertTrue(summary_path.exists())
            self.assertFalse((model_dir / "low_vol" / "model.txt").exists())
            self.assertIn("combined_test_metrics", summary)

    def test_train_dual_regime_models_reports_empty_regime_split_counts(self) -> None:
        def _make_split(start: str, regime_label: int, n: int) -> pd.DataFrame:
            idx = pd.date_range(start, periods=n, freq="min")
            base = np.linspace(-1.0, 1.0, n) + regime_label * 0.2
            future_return = base * 0.01
            return pd.DataFrame(
                {
                    "TDATE": idx,
                    "TRADE_DATE": idx.normalize(),
                    "REGIME_LABEL": regime_label,
                    "future_return": future_return,
                    "target_vol_scale": np.ones(n),
                    "target_vol_norm": future_return,
                    "f1": base,
                }
            )

        train_df = pd.concat([_make_split("2024-01-01 09:00:00", -1, 24), _make_split("2024-01-02 09:00:00", 1, 24)], ignore_index=True)
        val_df = _make_split("2024-01-03 09:00:00", -1, 12)
        test_df = pd.concat([_make_split("2024-01-04 09:00:00", -1, 12), _make_split("2024-01-05 09:00:00", 1, 8)], ignore_index=True)
        full_df = pd.concat([train_df, val_df, test_df], ignore_index=True)

        prepared = PreparedData(
            full_data=full_df,
            train_data=train_df,
            val_data=val_df,
            test_data=test_df,
            feature_cols=["f1"],
            factor_cols=["f1"],
            engineered_cols=[],
            runtime_factor_cols=["f1"],
            mid_weekly_cols=[],
            target_col="target_vol_norm",
            metadata={"feature_count": 1},
            feature_manifest={"all_feature_cols": ["f1"]},
        )

        with self.assertRaises(ValueError) as ctx:
            train_dual_regime_models(
                prepared=prepared,
                config_override={
                    "model": {
                        "persist_models": False,
                    },
                },
            )

        message = str(ctx.exception)
        self.assertIn("Regime 'high_vol' has empty split(s): val", message)
        self.assertIn("rows_by_split=train:24, val:0, test:8", message)

    def test_load_or_build_feature_frame_falls_back_when_cache_read_times_out(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            raw_path = root / "RBZL.SHF.csv"
            merged_cache = root / "cache.parquet"
            cache_meta = root / "cache_meta.json"
            regime_plot = root / "regime_plot.png"

            raw_path.write_text("TDATE,OPEN,HIGH,LOW,CLOSE\n2024-01-01 09:00:00,1,2,1,1.5\n", encoding="utf-8")
            merged_cache.write_bytes(b"placeholder")

            builder = FactorDatasetBuilder(
                config_override={
                    "paths": {
                        "raw_data": str(raw_path),
                        "merged_cache": str(merged_cache),
                        "cache_meta": str(cache_meta),
                        "regime_plot": str(regime_plot),
                    },
                }
            )

            fallback_result = (
                pd.DataFrame({"TDATE": pd.to_datetime(["2024-01-01 09:00:00"]), "TRADE_DATE": pd.to_datetime(["2024-01-01"])}),
                ["f1"],
                ["ENG_X"],
                ["f1"],
                [],
                {"factor_count": 1},
            )

            with patch("pipeline.dataset.pd.read_parquet", side_effect=TimeoutError(60, "Operation timed out")):
                with patch.object(builder, "build_feature_frame", return_value=fallback_result) as mocked_build_feature_frame:
                    result = builder.load_or_build_feature_frame(force_rebuild=False)

            mocked_build_feature_frame.assert_called_once()
            self.assertIs(result, fallback_result)


if __name__ == "__main__":
    unittest.main()
